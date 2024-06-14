import logging

import openpyxl

logger = logging.getLogger(__name__)


EXTENSIONS = "Extensions are additions"
ADDITONAL_FIELDS = "If you have additional information "
MAPPINGS_SHEETS = [
    "(OCDS) 1. General (all stages)",
    "(OCDS) 2. Planning",
    "(OCDS) 3. Tender",
    "(OCDS) 4. Award",
    "(OCDS) 5. Contract",
    "(OCDS) 6. Implementation",
]
DATA_SHEET = "2. Data Elements"
SCHEMA_SHEET = "OCDS Schema"


class Mapping:
    def __init__(self, config):
        self.config = config
        self.wb = openpyxl.load_workbook(self.config.file, data_only=True)
        self.data_elements = self.read_data_elements_sheet(self.wb[DATA_SHEET])
        # self.data_sources = self.read_sheet('1. Data Sources', skiprows=3)
        mappings = self.read_mappings()
        mappings = self.normmalize_mapping_column(mappings)
        self.mappings = self.enforce_mapping_structure(mappings)
        self.schema = self.read_schema_sheet()

    def get_schema_sheet(self):
        for sheet in self.wb.sheetnames:
            if SCHEMA_SHEET in sheet:
                return self.wb[sheet]
        return None

    def normmalize_mapping_column(self, mappings):
        """Normalize the mapping column by setting all space separators to one space"""
        for mapping in mappings:
            if "  " in mapping["mapping"]:
                mapping["mapping"] = " ".join((p.strip() for p in mapping["mapping"].split("  ")))
        return mappings

    def read_mapping_sheet(self, sheet):
        headers = [cell.value for cell in sheet[3]]
        # Iterate over the rows, starting from the third row
        in_extensions = False
        current_extension = ""
        mappings = []

        for row in sheet.iter_rows(min_row=4, values_only=True):
            if len(row) != len(headers):
                logger.debug(f"Row {row} does not have the same number of columns as the headers")
                continue
            column_type = row[0]
            path = row[2]
            title = row[3]
            description = row[4]
            mapping = row[5]
            match column_type:
                case "field" | "required_field" | "extension_field" | "additional_field":
                    if not mapping:
                        continue
                    row = {
                        "path": path,
                        "title": title,
                        "description": description,
                        "mapping": mapping,
                        "is_extensions": in_extensions,
                        "extension": current_extension,
                        "is_required": column_type == "required_field",
                        "is_additional": column_type == "additional_field",
                    }
                    mappings.append(row)
                case "span" | "extension_span" | "subtitle" | "ref_span" | "required_span":
                    # informational, but not useful for machine
                    continue
                case "section":
                    if EXTENSIONS in path or ADDITONAL_FIELDS in path:
                        in_extensions = True
                case "extension":
                    current_extension = path.split(":")[0]
        return mappings

    def read_mappings(self):
        mappings = []
        for sheet_name in MAPPINGS_SHEETS:
            sheet = self.wb[sheet_name]
            mappings.extend(self.read_mapping_sheet(sheet))
        return mappings

    def get_element_by_mapping(self, for_mapping):
        element = for_mapping.split("(")[1].replace(")", "")
        return self.data_elements.get(element)

    def read_data_elements_sheet(self, sheet):
        elements = {}
        for row in sheet.iter_rows(min_row=4, values_only=True):
            for_mapping, data_source, table, data_element, publish, example, description, data_type, _, _, _, _ = row
            if not data_element:
                continue
            elements[data_element] = {
                "for_mapping": str(for_mapping).replace("  ", " "),
                "data_source": data_source,
                "data_element": data_element,
                "table": table,
                "publish": "yes" in str(publish).lower(),
                "example": example,
                "description": description,
                "data_type": data_type,
            }
        return elements

    def read_schema_sheet(self):
        sheet = self.get_schema_sheet()
        if sheet is None:
            return {}
        schema = {}
        for row in sheet.iter_rows(min_row=2, values_only=True):
            _, path, title, description, type, range, values, links, _, _ = row
            if not path:
                continue
            path = "/" + path

            if path in schema and type == "object":
                # this is a nested object inside arrray we are interested in parrent array path
                continue
            schema[path] = {
                "title": title,
                "description": description,
                "type": type,
                "range": range,
                "values": values,
                "links": links,
            }
        return schema

    def enforce_mapping_structure(self, mappings):
        sections = {"planning": [], "tender": [], "awards": [], "contracts": [], "implementation": [], "general": []}
        for mapping in mappings:
            section = mapping["path"].split("/")[0]
            if section in sections:
                sections[section].append(mapping)
            else:
                sections["general"].append(mapping)
        # Merge the lists in the specified order
        return (
            sections["general"]
            + sections["planning"]
            + sections["tender"]
            + sections["awards"]
            + sections["contracts"]
            + sections["implementation"]
        )

    def get_mapping(self):
        return self.mappings

    def get_mapping_for(self, path):
        result = []
        for mapping in self.mappings:
            if mapping["path"] == path:
                result.append(mapping)
        return result

    # def get_data_sources(self):
    #     return self.data_sources.asdict()

    def get_paths_for_mapping(self, key, force_publish=False):
        result = []
        for mapping in self.mappings:
            if mapping["mapping"] == key:
                element = self.get_element_by_mapping(key)
                if element["publish"] or force_publish:
                    result.append(mapping["path"])
        return result

    def is_array_path(self, path):
        return self.schema.get(path, {}).get("type") == "array"

    def get_arrays(self):
        result = []
        for path, schema in self.schema.items():
            if schema["type"] == "array":
                result.append(path)
        return result

    def get_scheme(self):
        return self.schema

    def get_ocid_mapping(self):
        ocid_mapping = self.get_mapping_for("ocid")[0]
        return ocid_mapping["mapping"]
