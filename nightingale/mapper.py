import openpyxl
import tomllib
import flatten_dict
import dict_hash
import simplejson as json
from datetime import datetime
from collections import defaultdict
from pathlib import Path
from typing import Any
from pdb import set_trace
from nightingale.transformer import DataTransformer
#from .readers import READERS
from .config.v1 import TransformationConfig


def read_workbook(path):
    data = defaultdict(list)
    wb = openpyxl.load_workbook(path)
    for sheet in wb.sheetnames:
        sheet_data = wb[sheet]
        headers = [cell.value for cell in sheet_data[1]]
        for row in sheet_data.iter_rows(min_row=2):
            row_dict = {header: cell.value for header, cell in zip(headers, row)}
            data[sheet].append(row_dict)
    return data


READERS = {
    'xlsx': read_workbook,
    'csv': openpyxl.load_workbook,
}

class MapperConfig:

    def __init__(self, path: Path):
        with path.open('rb') as f:
            self._config = tomllib.load(f)
        self.base = self._config['datasources'].pop('directory')

    def get_ocid_prefix(self):
        return self._config['mapper']['ocid-prefix']

    def get_publisher(self):
        return self._config['mapper']['publisher']

    def get_ocds_version(self):
        return self._config['mapper']['version']

    def get_map_conf(self) -> dict:
        return self._config['mapper']

    def get_output_dir(self):
        return self._config['output']['directory']

    def get_sheet_names(self) -> dict[str, Any]:
        return self._config['mapper']['sheets']

    def get_input_file(self):
        return self._config['source']['input_file']

    def _full_path(self, f):
        return Path(self.base)/ f

    def get_datasources(self):
        for k, v in self._config['datasources'].items():
            files = [self._full_path(f) for f in v['files']]
            v['files'] = files
            yield k, v

    def get_datasource_info(self, name):
        ds = self._config['datasources'].get(name, {})
        if ds:
            files = [self._full_path(f) for f in ds['files']]
            ds['files'] = files
        return ds


class OCDSDataMapper:

    def __init__(self, config: MapperConfig):
        self.config = config
        self.tr_conf = TransformationConfig(self.config)

    def get_data_sources(self):
        data_sources = dict(self.config.get_datasources())
        for name, datsource in self.tr_conf.get_datasources():
            if name in data_sources:
           # 3000+ number contract number to join system
           #
                yield (name, datsource)

    def transform(self):
        current_ocid = ''
        data = {}
        for name, datasource in self.get_data_sources():
            print(f'Mapping datasource {name}')
            conf = self.config.get_datasource_info(name.strip())
            rows = []
            for file in conf['files']:
                reader = READERS[conf['format']]
                data = reader(file)
                for sheet_name, sheet_data in reader(file).items():
                    mapping  = self.tr_conf.get_mapping_for_sheet(sheet_name)
                    elements = dict(iter(self.tr_conf.get_data_elements(name, sheet_name)))
                    ocid_mapping = mapping.pop('ocid')
                    ocid_col = elements[ocid_mapping['Mapping']]
                    for i, row in enumerate(sheet_data):
                        if i > 10:
                            break
                        data = {}
                        ocid = self.produce_ocid(row[ocid_col['Data element']])
                        data[ocid_mapping['Path']] = ocid
                        for map in mapping.values():
                            column = elements[map['Mapping']]
                            value = row.get(column['Data element'])
                            data[map['Path']] = value
                        id_ = dict_hash.sha256(data)
                        data['id'] = id_
                        rows.append(flatten_dict.unflatten(data, splitter='path'))
            package  = self.produce_empty_package()
            package['releases'] = rows
            self.write_package(package)


    def produce_ocid(self, value):
        prefix = self.config.get_ocid_prefix()
        return f'{prefix}-{value}'

    def produce_empty_package(self):
        now = datetime.now().isoformat()
        return {
            'uri': f'https://todo.com/{now}',
            'version': self.config.get_ocds_version(),
            'publisher': self.config.get_publisher(),
            'publishedDate': now,
            'releases': []
        }

    def write_package(self, package):
        base = Path(self.config.get_output_dir())
        base.mkdir(parents=True, exist_ok=True)
        base = base / f'release-package-{package['publishedDate']}.json'
        with  base.open('w') as f:
            json.dump(package, f, indent=2)

    def _transform_general_stage(self, data):
        pass
    def _trasform_stage(self, data):
        pass


# conf = Path('test.toml')
# tr = OCDSDataMapper(MapperConfig(conf))
# tr.transform()
