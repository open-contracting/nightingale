import logging

import openpyxl

from nightingale.utils import get_longest_array_path

logger = logging.getLogger(__name__)


EXTENSIONS = "Extensions are additions"
ADDITONAL_FIELDS = "If you have additional information "
MAPPINGS_SHEETS = [
    "(OCDS) 1. General (all stages)",
    "(OCDS) 2. Planning",
    "(OCDS) 3. Tender",
    "(OCDS) 4. Award",
    "(OCDS) 5. Contract",
    "(OCDS) 6. Implementation",
]
DATA_SHEET = "2. Data Elements"
EXTENSIONS_SHEET = "3. OCDS Extensions"
SCHEMA_SHEET = "Schema"


class MappingTemplate:
    def __init__(self, config):
        self.config = config
        self.wb = openpyxl.load_workbook(self.config.file, data_only=True)
        self.data_elements = self.read_data_elements_sheet(self.wb[DATA_SHEET])
        # self.data_sources = self.read_sheet('1. Data Sources', skiprows=3)
        mappings = self.read_mappings()
        # mappings = self.normmalize_mapping_column(mappings)
        self.mappings = self.enforce_mapping_structure(mappings)
        self.schema = self.read_schema_sheet()
        self.extensions = self.read_extenions_info()

    def get_schema_sheet(self):
        return [self.wb[sheet] for sheet in self.wb.sheetnames if "OCDS" in sheet and SCHEMA_SHEET in sheet]

    def normmalize_mapping_column(self, mappings):
        """Normalize the mapping column by setting all space separators to one space"""
        for mapping in mappings:
            if "  " in mapping["mapping"]:
                mapping["mapping"] = " ".join((p.strip() for p in mapping["mapping"].split("  ")))
        return mappings

    def read_mapping_sheet(self, sheet):
        # Iterate over the rows, starting from the third row
        in_extensions = False
        current_extension = ""
        current_block = ""
        mappings = []

        for row in sheet.iter_rows(min_row=4, values_only=True):
            column_type = row[0]
            path = row[2]
            title = row[3]
            description = row[4]
            mapping = row[5]
            match column_type:
                case "span" | "ref_span" | "extension_span":
                    current_block = path
                case "field" | "required_field" | "extension_field" | "additional_field":
                    if not mapping:
                        continue
                    row = {
                        "block": current_block,
                        "path": path if path.startswith("/") else "/" + path,
                        "title": title,
                        "description": description,
                        "mapping": mapping,
                        "is_extensions": in_extensions,
                        "extension": current_extension,
                        "is_required": column_type == "required_field",
                        "is_additional": column_type == "additional_field",
                    }
                    mappings.append(row)
                case "subtitle" | "required_span":
                    continue
                case "section":
                    if EXTENSIONS in path or ADDITONAL_FIELDS in path:
                        in_extensions = True
                case "extension":
                    current_extension = path.split(":")[0]
        return mappings

    def read_mappings(self):
        mappings = []
        for sheet_name in MAPPINGS_SHEETS:
            sheet = self.wb[sheet_name]
            mappings.extend(self.read_mapping_sheet(sheet))
        return mappings

    def get_element_by_mapping(self, for_mapping):
        element = for_mapping.split("(")[1].replace(")", "")
        return self.data_elements.get(element, {})

    def read_data_elements_sheet(self, sheet):
        elements = {}
        for row in sheet.iter_rows(min_row=4, values_only=True):
            for_mapping, data_source, table, data_element, publish, example, description, data_type, *_ = row
            if not data_element:
                continue
            elements[data_element] = {
                "for_mapping": str(for_mapping).replace("  ", " "),
                "data_source": data_source,
                "data_element": data_element,
                "table": table,
                "publish": "yes" in str(publish).lower(),
                "example": example,
                "description": description,
                "data_type": data_type,
            }
        return elements

    def read_extenions_info(self):
        if EXTENSIONS_SHEET not in self.wb.sheetnames:
            return []
        sheet = self.wb[EXTENSIONS_SHEET]
        headers = [(i, cell.value) for i, cell in enumerate(sheet[1])]  # Assuming the first row contains the headers
        data = []
        for row in sheet.iter_rows(min_row=2, values_only=True):
            if not any(row):
                continue
            row_data = {k.lower(): row[i] for i, k in headers if k}
            data.append(row_data)
        return data

    def get_data_elements(self):
        return self.data_elements

    def read_schema_sheet(self):
        sheets = self.get_schema_sheet()
        if sheets is None:
            return {}
        schema = {}
        for sheet in sheets:
            for row in sheet.iter_rows(min_row=2, values_only=True):
                _, path, title, description, type, range, values, links, codelist, *_ = row
                if not path:
                    continue
                path = "/" + path

                if path in schema and type == "object":
                    # this is a nested object inside arrray we are interested in parrent array path
                    continue
                schema[path] = {
                    "title": title,
                    "description": description,
                    "type": "array" if type and "array" in type.lower() else type,
                    "range": range,
                    "values": values,
                    "links": links,
                    "codelist": codelist,
                }
        return schema

    def enforce_mapping_structure(self, mappings):
        sections = {"planning": [], "tender": [], "awards": [], "contracts": [], "implementation": [], "general": []}
        for mapping in mappings:
            section = mapping["path"].split("/")[0]
            if section in sections:
                sections[section].append(mapping)
            else:
                sections["general"].append(mapping)
        # Merge the lists in the specified order
        return (
            sections["general"]
            + sections["planning"]
            + sections["tender"]
            + sections["awards"]
            + sections["contracts"]
            + sections["implementation"]
        )

    def get_mappings(self):
        return self.mappings

    def get_mapping_for(self, path):
        result = []
        if not path.startswith("/"):
            path = "/" + path
        for mapping in self.get_mappings():
            if mapping["path"] == path:
                result.append(mapping)
        return result

    # def get_data_sources(self):
    #     return self.data_sources.asdict()

    def get_paths_for_mapping(self, key, force_publish=False):
        result = []
        for mapping in self.get_mappings():
            if mapping["mapping"] == key:
                element = self.get_element_by_mapping(key)
                if element.get("publish", False) or force_publish:
                    result.append(mapping["path"])
        return result

    def is_array_path(self, path):
        return self.schema.get(path, {}).get("type") == "array"

    def get_arrays(self):
        result = []
        for path, schema in self.schema.items():
            if schema["type"] == "array":
                result.append(path)
        return result

    def get_schema(self):
        return self.schema

    def get_ocid_mapping(self):
        ocid_mapping = self.get_mapping_for("ocid")[0]
        return ocid_mapping["mapping"]

    def get_containing_array_path(self, path):
        return get_longest_array_path(self.get_arrays(), path)

    def get_datetime_fields(self):
        """
        Returns a list of paths that are marked as 'date-time' in the 'values' column in the schema.
        """
        datetime_fields = []
        for path, field_info in self.schema.items():
            if field_info.get("values") == "date-time":
                datetime_fields.append(path)
        return datetime_fields
