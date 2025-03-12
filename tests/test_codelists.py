import openpyxl
import pytest

from nightingale.codelists import CodelistsMapping


# Dummy classes to simulate an openpyxl workbook and sheet
class DummySheet:
    def __init__(self, title, rows):
        self.title = title
        self._rows = rows

    def iter_rows(self, values_only=True):
        return iter(self._rows)


class DummyWorkbook:
    def __init__(self, sheets):
        self.worksheets = sheets


# Dummy config to pass to CodelistsMapping
class DummyConfig:
    pass


@pytest.fixture
def dummy_config(monkeypatch):
    cfg = DummyConfig()
    cfg.codelists = "dummy.xlsx"
    # Return an empty workbook to avoid actual file I/O
    monkeypatch.setattr(openpyxl, "load_workbook", lambda filename, data_only: DummyWorkbook([]))
    return cfg


def test_normmalize_mapping_column(dummy_config):
    # Instantiate CodelistsMapping (workbook is empty in this case)
    cm = CodelistsMapping(dummy_config)
    mappings = [
        {"mapping": "A  B  C"},
        {"mapping": "NoExtraSpaces"},
        {"mapping": "X   Y"},  # three spaces becomes one space each
    ]
    normalized = cm.normmalize_mapping_column(mappings)
    assert normalized[0]["mapping"] == "A B C"
    assert normalized[1]["mapping"] == "NoExtraSpaces"
    assert normalized[2]["mapping"] == "X Y"


def test_read_codelists_sheet(dummy_config):
    # Create a dummy sheet that simulates a codelist mapping.
    # All rows now contain four items.
    rows = [
        ("codelist_name", "codelist: MyList", None, None),
        ("codelist_headers", "Code", "Source codelist", "Source code"),
        ("A", "A", "Src", "X"),
        ("B", "B", "Src", None),  # should be skipped because source code is missing
        ("C", "C", "Src", "Y"),
    ]
    sheet = DummySheet("(OCDS) Test", rows)
    cm = CodelistsMapping(dummy_config)
    mapping = cm.read_codelists_sheet(sheet)
    # Expected: the codelist "MyList" with mappings for "X" and "Y"
    expected = {"MyList": {"X": "A", "Y": "C"}}
    assert mapping == expected


def test_load_codelists_mapping(dummy_config, monkeypatch):
    # Create two dummy sheets; only the one with title starting with "(ocds)" should be processed.
    rows_sheet1 = [
        ("codelist_name", "codelist: List1", None, None),
        ("codelist_headers", "Code", "Source codelist", "Source code"),
        ("10", "10", "S1", "val1"),
    ]
    rows_sheet2 = [
        ("codelist_name", "codelist: List2", None, None),
        ("codelist_headers", "Code", "Source codelist", "Source code"),
        ("20", "20", "S2", "val2"),
    ]
    sheet1 = DummySheet("(OCDS) Sheet1", rows_sheet1)
    sheet2 = DummySheet("Other Sheet", rows_sheet2)
    dummy_wb = DummyWorkbook([sheet1, sheet2])
    monkeypatch.setattr(openpyxl, "load_workbook", lambda filename, data_only: dummy_wb)
    cm = CodelistsMapping(dummy_config)
    # Only sheet1 is processed, so we expect mapping from List1 only.
    expected = {"List1": {"val1": "10"}}
    assert cm.codelists == expected


def test_get_mapping_for_codelist(dummy_config, monkeypatch):
    # Create a dummy sheet with a codelist mapping.
    rows = [
        ("codelist_name", "codelist: TestCodelist", None, None),
        ("codelist_headers", "Code", "Source codelist", "Source code"),
        ("1", "1", "TestSource", "ABC"),
        ("2", "2", "TestSource", None),  # missing source code; should be skipped
    ]
    dummy_sheet = DummySheet("(OCDS) Test", rows)
    dummy_wb = DummyWorkbook([dummy_sheet])
    monkeypatch.setattr(openpyxl, "load_workbook", lambda filename, data_only: dummy_wb)
    cm = CodelistsMapping(dummy_config)
    mapping = cm.get_mapping_for_codelist("TestCodelist")
    assert mapping == {"ABC": "1"}


if __name__ == "__main__":
    pytest.main()
