from unittest.mock import MagicMock, patch

import openpyxl
import pytest

from nightingale.mapping_template.v09 import MappingTemplate
from nightingale.utils import get_longest_array_path


def _sheets():
    return {
        "OCDS Schema": [
            (None, "array_path", "title1", "description1", "array", "range1", "values1", "links1", None, None),
            (None, "path1", "title2", "description2", "string", "range2", "values2", "links2", None, None),
            (None, "", "not set", "not set", "string", "range2", "values2", "links2", None, None),
            (
                None,
                "array_path",
                "object_inside_array",
                "object_inside_array",
                "object",
                "",
                "values2",
                "links2",
                None,
                None,
            ),
        ],
        "2. Data Elements": [
            (
                "mapping1",
                "source1",
                "table1",
                "element1",
                "yes",
                "example1",
                "description1",
                "type1",
                None,
                None,
                None,
                None,
            ),
            (
                "mapping2",
                "source2",
                "table2",
                "element2",
                "no",
                "example2",
                "description2",
                "type2",
                None,
                None,
                None,
                None,
            ),
            (
                "mapping3",
                "source3",
                "table3",
                "",
                "no",
                "not set",
                "not set",
                "type2",
                None,
                None,
                None,
                None,
            ),
        ],
        "ocds": [
            ("field", "id", "path", "title", "description", "mapping", None, None, None, None),
            ("required_field", "id", "path2", "title2", "description2", "mapping2", None, None, None, None),
            ("field", "id", "path_not_mapped", "title_not_mapped", "description", "", None, None, None, None),
            ("span", "text", "text", "text", "description", "", None, None, None, None),
            ("section", "text", "Extensions are additions", "text", "description", "", None, None, None, None),
            ("extension", "extension_name", "extension:description", "", "", "", None, None, None, None),
            (
                "additional_field",
                "extension_field",
                "extension_field",
                "extension_field_title",
                "description",
                "extension_mapping",
                None,
                None,
                None,
                None,
            ),
        ],
    }


def getter(key):
    sheets = _sheets()
    mock_sheet = MagicMock()
    if "schema" in key.lower():
        data = sheets["OCDS Schema"]
    elif "elements" in key.lower():
        data = sheets["2. Data Elements"]
    elif "ocds" in key.lower():
        data = sheets["ocds"]
        mock_sheet.__getitem__.side_effect = lambda name: [MagicMock() for _ in range(10)]
    mock_sheet.iter_rows.return_value = data
    return mock_sheet


@pytest.fixture
def mock_workbook():
    workbook = MagicMock(spec=openpyxl.Workbook)
    workbook.sheetnames = [
        "(OCDS) 1. General (all stages)",
        "(OCDS) 2. Planning",
        "(OCDS) 3. Tender",
        "(OCDS) 4. Award",
        "(OCDS) 5. Contract",
        "(OCDS) 6. Implementation",
        "2. Data Elements",
        "OCDS Schema 1.1.5",
    ]
    # Mocking the sheets in the workbook to be accessed via subscript notation
    workbook.__getitem__.side_effect = lambda name: MagicMock()
    return workbook


@pytest.fixture
def mock_config():
    class Config:
        file = "/mnt/data/mapping.xlsx"

    return Config()


@patch("openpyxl.load_workbook")
def test_mapping_init(mock_load_workbook, mock_workbook, mock_config):
    mock_load_workbook.return_value = mock_workbook
    mapping = MappingTemplate(mock_config)
    assert mapping.wb == mock_workbook


@patch("openpyxl.load_workbook")
def test_normmalize_mapping_column(mock_config):
    mappings = [
        {"mapping": "field1  field2"},
        {"mapping": "field3   field4"},
    ]
    expected_result = [
        {"mapping": "field1 field2"},
        {"mapping": "field3 field4"},
    ]
    mapping = MappingTemplate(mock_config)
    result = mapping.normmalize_mapping_column(mappings)
    assert result == expected_result


@patch("openpyxl.load_workbook")
def test_read_data_elements_sheet(mock_load_workbook, mock_workbook, mock_config):
    mock_workbook.__getitem__ = lambda self, x: getter(x)
    mock_load_workbook.return_value = mock_workbook

    mapping = MappingTemplate(mock_config)

    expected_data_elements = {
        "element1": {
            "for_mapping": "mapping1",
            "data_source": "source1",
            "data_element": "element1",
            "table": "table1",
            "publish": True,
            "example": "example1",
            "description": "description1",
            "data_type": "type1",
        },
        "element2": {
            "for_mapping": "mapping2",
            "data_source": "source2",
            "data_element": "element2",
            "table": "table2",
            "publish": False,
            "example": "example2",
            "description": "description2",
            "data_type": "type2",
        },
    }

    assert mapping.data_elements == expected_data_elements


@patch("openpyxl.load_workbook")
def test_read_schema_sheet(mock_load_workbook, mock_workbook, mock_config):
    mock_workbook.__getitem__ = lambda self, x: getter(x)
    mock_load_workbook.return_value = mock_workbook

    mapping = MappingTemplate(mock_config)
    schema = mapping.get_schema()

    expected_schema = {
        "/array_path": {
            "codelist": None,
            "title": "title1",
            "description": "description1",
            "type": "array",
            "range": "range1",
            "values": "values1",
            "links": "links1",
        },
        "/path1": {
            "codelist": None,
            "title": "title2",
            "description": "description2",
            "type": "string",
            "range": "range2",
            "values": "values2",
            "links": "links2",
        },
    }
    assert schema == expected_schema
    assert mapping.get_arrays() == ["/array_path"]
    assert mapping.get_containing_array_path("/array_path/id") == "/array_path"


@patch("openpyxl.load_workbook")
def test_enforce_mapping_structure(mock_load_workbook, mock_workbook, mock_config):
    mock_load_workbook.return_value = mock_workbook

    mappings = [
        {"path": "planning/item1", "mapping": "map1"},
        {"path": "tender/item2", "mapping": "map2"},
        {"path": "awards/item3", "mapping": "map3"},
        {"path": "contracts/item4", "mapping": "map4"},
        {"path": "implementation/item5", "mapping": "map5"},
        {"path": "general/item6", "mapping": "map6"},
    ]

    mapping = MappingTemplate(mock_config)
    result = mapping.enforce_mapping_structure(mappings)

    expected_result = [
        {"path": "general/item6", "mapping": "map6"},
        {"path": "planning/item1", "mapping": "map1"},
        {"path": "tender/item2", "mapping": "map2"},
        {"path": "awards/item3", "mapping": "map3"},
        {"path": "contracts/item4", "mapping": "map4"},
        {"path": "implementation/item5", "mapping": "map5"},
    ]

    assert result == expected_result


@patch("openpyxl.load_workbook")
def test_read_mappings(mock_load_workbook, mock_workbook, mock_config):
    mock_workbook.__getitem__ = lambda self, x: getter(x)
    mock_load_workbook.return_value = mock_workbook

    mapping = MappingTemplate(mock_config)

    expected_mappings = [
        {
            "path": "/path",
            "title": "title",
            "description": "description",
            "mapping": "mapping",
            "is_extensions": False,
            "extension": "",
            "is_required": False,
            "is_additional": False,
        },
        {
            "path": "/path2",
            "title": "title2",
            "description": "description2",
            "mapping": "mapping2",
            "is_extensions": False,
            "extension": "",
            "is_required": True,
            "is_additional": False,
        },
        {
            "description": "description",
            "extension": "extension",
            "is_additional": True,
            "is_extensions": True,
            "is_required": False,
            "mapping": "extension_mapping",
            "path": "/extension_field",
            "title": "extension_field_title",
        },
        {
            "path": "/path",
            "title": "title",
            "description": "description",
            "mapping": "mapping",
            "is_extensions": False,
            "extension": "",
            "is_required": False,
            "is_additional": False,
        },
        {
            "path": "/path2",
            "title": "title2",
            "description": "description2",
            "mapping": "mapping2",
            "is_extensions": False,
            "extension": "",
            "is_required": True,
            "is_additional": False,
        },
        {
            "description": "description",
            "extension": "extension",
            "is_additional": True,
            "is_extensions": True,
            "is_required": False,
            "mapping": "extension_mapping",
            "path": "/extension_field",
            "title": "extension_field_title",
        },
        {
            "path": "/path",
            "title": "title",
            "description": "description",
            "mapping": "mapping",
            "is_extensions": False,
            "extension": "",
            "is_required": False,
            "is_additional": False,
        },
        {
            "path": "/path2",
            "title": "title2",
            "description": "description2",
            "mapping": "mapping2",
            "is_extensions": False,
            "extension": "",
            "is_required": True,
            "is_additional": False,
        },
        {
            "description": "description",
            "extension": "extension",
            "is_additional": True,
            "is_extensions": True,
            "is_required": False,
            "mapping": "extension_mapping",
            "path": "/extension_field",
            "title": "extension_field_title",
        },
        {
            "path": "/path",
            "title": "title",
            "description": "description",
            "mapping": "mapping",
            "is_extensions": False,
            "extension": "",
            "is_required": False,
            "is_additional": False,
        },
        {
            "path": "/path2",
            "title": "title2",
            "description": "description2",
            "mapping": "mapping2",
            "is_extensions": False,
            "extension": "",
            "is_required": True,
            "is_additional": False,
        },
        {
            "description": "description",
            "extension": "extension",
            "is_additional": True,
            "is_extensions": True,
            "is_required": False,
            "mapping": "extension_mapping",
            "path": "/extension_field",
            "title": "extension_field_title",
        },
        {
            "path": "/path",
            "title": "title",
            "description": "description",
            "mapping": "mapping",
            "is_extensions": False,
            "extension": "",
            "is_required": False,
            "is_additional": False,
        },
        {
            "path": "/path2",
            "title": "title2",
            "description": "description2",
            "mapping": "mapping2",
            "is_extensions": False,
            "extension": "",
            "is_required": True,
            "is_additional": False,
        },
        {
            "description": "description",
            "extension": "extension",
            "is_additional": True,
            "is_extensions": True,
            "is_required": False,
            "mapping": "extension_mapping",
            "path": "/extension_field",
            "title": "extension_field_title",
        },
        {
            "path": "/path",
            "title": "title",
            "description": "description",
            "mapping": "mapping",
            "is_extensions": False,
            "extension": "",
            "is_required": False,
            "is_additional": False,
        },
        {
            "path": "/path2",
            "title": "title2",
            "description": "description2",
            "mapping": "mapping2",
            "is_extensions": False,
            "extension": "",
            "is_required": True,
            "is_additional": False,
        },
        {
            "description": "description",
            "extension": "extension",
            "is_additional": True,
            "is_extensions": True,
            "is_required": False,
            "mapping": "extension_mapping",
            "path": "/extension_field",
            "title": "extension_field_title",
        },
    ]
    mapping.get_mappings()
    assert mapping.get_mappings() == expected_mappings
    assert mapping.get_mapping_for("path2") == [
        {
            "path": "/path2",
            "title": "title2",
            "description": "description2",
            "mapping": "mapping2",
            "is_extensions": False,
            "extension": "",
            "is_required": True,
            "is_additional": False,
        }
        for _ in range(6)
    ]


@patch("openpyxl.load_workbook")
def test_get_element_by_mapping(mock_load_workbook, mock_workbook, mock_config):
    mock_sheet = MagicMock()
    mock_sheet.iter_rows.return_value = iter(
        [
            (
                "mapping1",
                "source1",
                "table1",
                "element1",
                "yes",
                "example1",
                "description1",
                "type1",
                None,
                None,
                None,
                None,
            ),
            (
                "mapping2",
                "source2",
                "table2",
                "element2",
                "no",
                "example2",
                "description2",
                "type2",
                None,
                None,
                None,
                None,
            ),
        ]
    )
    mock_workbook.__getitem__.side_effect = lambda name: mock_sheet
    mock_load_workbook.return_value = mock_workbook

    mapping = MappingTemplate(mock_config)
    element = mapping.get_element_by_mapping("mapping1 (element1)")

    expected_element = {
        "for_mapping": "mapping1",
        "data_source": "source1",
        "data_element": "element1",
        "table": "table1",
        "publish": True,
        "example": "example1",
        "description": "description1",
        "data_type": "type1",
    }

    assert element == expected_element


@patch("openpyxl.load_workbook")
def test_get_paths_for_mapping(mock_load_workbook, mock_workbook, mock_config):
    mock_sheet = MagicMock()
    mock_sheet.iter_rows.return_value = iter(
        [
            (
                "mapping1",
                "source1",
                "table1",
                "element1",
                "yes",
                "example1",
                "description1",
                "type1",
                None,
                None,
                None,
                None,
            ),
            (
                "mapping2",
                "source2",
                "table2",
                "element2",
                "no",
                "example2",
                "description2",
                "type2",
                None,
                None,
                None,
                None,
            ),
        ]
    )
    mock_workbook.__getitem__.side_effect = lambda name: mock_sheet
    mock_load_workbook.return_value = mock_workbook

    mock_mappings = [
        {
            "path": "path1",
            "mapping": "mapping1 (element1)",
            "is_extensions": False,
            "extension": "",
            "is_required": False,
            "is_additional": False,
        },
        {
            "path": "path2",
            "mapping": "mapping2 (element2)",
            "is_extensions": False,
            "extension": "",
            "is_required": False,
            "is_additional": False,
        },
    ]

    mapping = MappingTemplate(mock_config)
    mapping.mappings = mock_mappings

    paths = mapping.get_paths_for_mapping("mapping1 (element1)")

    expected_paths = ["path1"]

    assert paths == expected_paths


@patch("openpyxl.load_workbook")
def test_is_array_path(mock_load_workbook, mock_workbook, mock_config):
    mock_workbook.__getitem__ = lambda self, x: getter(x)
    mock_load_workbook.return_value = mock_workbook

    mapping = MappingTemplate(mock_config)
    mapping.schema = mapping.read_schema_sheet()
    assert mapping.is_array_path("/array_path") is True
    assert mapping.is_array_path("/path2") is False


def test_get_longest_array_path():
    arrays = ["/root/level1/level2", "/root/level1", "/root/level1/level2/level3"]
    path = "/root/level1/level2/element"
    result = get_longest_array_path(arrays, path)
    assert result == "/root/level1/level2"


if __name__ == "__main__":
    pytest.main()
